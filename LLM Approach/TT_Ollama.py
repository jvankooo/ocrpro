import fitz  # PyMuPDF
import pdfplumber
import pandas as pd
import pytesseract
import json
import os
import numpy as np
import cv2
from PIL import Image
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
from io import BytesIO
import layoutparser as lp
import torch
from transformers import TableTransformerForObjectDetection, DetrImageProcessor
import warnings

warnings.filterwarnings('ignore')

# Load Table Transformer model
device = "cuda" if torch.cuda.is_available() else "cpu"
processor = DetrImageProcessor.from_pretrained("microsoft/table-transformer-detection")
model = TableTransformerForObjectDetection.from_pretrained("microsoft/table-transformer-detection").to(device)


def extract_metadata(pdf_path):
    """Extract metadata from the PDF file."""
    doc = fitz.open(pdf_path)
    metadata = doc.metadata
    doc.close()
    return metadata


def extract_text_by_page(pdf_path, output_path="extracted_text.txt"):
    """Extract text directly from the PDF page by page, combining native text extraction with OCR."""
    doc = fitz.open(pdf_path)
    all_text = []
    
    for page_num, page in enumerate(doc):
        text = page.get_text("text")
        pix = page.get_pixmap(matrix=fitz.Matrix(2,2))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        ocr_text = pytesseract.image_to_string(img, config='--psm 6')
        combined_text = text.strip() + "\n" + ocr_text.strip()
        all_text.append(f"--- Page {page_num + 1} ---\n{combined_text}\n")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(all_text))
    doc.close()


def extract_images_and_diagrams(pdf_path, output_folder="extracted_images"):
    """Extract images and charts from PDF, filtering meaningful content."""
    os.makedirs(output_folder, exist_ok=True)
    doc = fitz.open(pdf_path)
    
    for page_num, page in enumerate(doc):
        for img_index, img in enumerate(page.get_images(full=True)):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image_ext = base_image["ext"]
            img = np.array(Image.open(BytesIO(image_bytes)))
            
            if np.std(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)) > 10:
                image_filename = f"{output_folder}/image_p{page_num+1}_{img_index+1}.{image_ext}"
                with open(image_filename, "wb") as img_file:
                    img_file.write(image_bytes)
    doc.close()


def detect_tables_with_table_transformer(image):
    """Detect tables in an image using Table Transformer model."""
    encoding = processor(images=image, return_tensors="pt").to(device)
    with torch.no_grad():
        outputs = model(**encoding)
    return outputs


def extract_tables(pdf_path):
    """Extract tables using Table Transformer and pdfplumber."""
    doc = fitz.open(pdf_path)
    tables = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            extracted_tables = page.extract_tables()
            for table in extracted_tables:
                df = pd.DataFrame(table)
                if df.shape[0] >= 2 and df.shape[1] >= 2:
                    tables.append((page_num+1, df))
    
    for page_num, page in enumerate(doc):
        pix = page.get_pixmap(matrix=fitz.Matrix(2,2))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        outputs = detect_tables_with_table_transformer(img)
        if len(outputs.logits) > 0:
            ocr_text = pytesseract.image_to_string(img, config='--psm 6')
            table_df = parse_text_to_table(ocr_text)
            if table_df.shape[0] >= 2 and table_df.shape[1] >= 2:
                tables.append((page_num+1, table_df))
    
    doc.close()
    return tables


def save_tables_to_excel(tables, output_file="extracted_tables.xlsx"):
    """Save tables to Excel with each table on its own sheet."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    for page_num, table in tables:
        sheet_name = f"Page_{page_num}_Table"
        ws = wb.create_sheet(title=sheet_name[:31])
        for row in dataframe_to_rows(table, index=False, header=True):
            ws.append(row)
    
    wb.save(output_file)


def process_pdf(pdf_path):
    metadata = extract_metadata(pdf_path)
    extract_text_by_page(pdf_path)
    extract_images_and_diagrams(pdf_path)
    tables = extract_tables(pdf_path)
    save_tables_to_excel(tables)
    print("Processing complete! Metadata:", metadata)


# Example Usage
pdf_path = r"E:\Btech_AI\Intern\ocrpro\Phable CAM Final.pdf"
process_pdf(pdf_path)
